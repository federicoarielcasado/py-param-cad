"""
BOMGenerator — Bill of Materials generation from revision data (Semana 11).

Public API:
    gen = BOMGenerator()
    result = gen.generate(piece_code, parameters, output_dir, ctx)

Output:
    - .xlsx (openpyxl) — styled spreadsheet with header block and data table
    - .pdf  (reportlab) — A4 landscape table with matching layout

Piece weight is computed from geometry + material density when available.
BOM items are returned as plain dicts so PieceController can persist them
via BOMRepository (generators never touch the DB directly).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional


# ---------------------------------------------------------------------------
# Public data classes
# ---------------------------------------------------------------------------

@dataclass
class BOMContext:
    """Metadata written into the BOM header block."""
    piece_code: str
    piece_display_name: str
    drawing_number: str
    revision_code: str
    author: str
    company_name: str
    date_str: str = ""

    def __post_init__(self) -> None:
        if not self.date_str:
            self.date_str = date.today().strftime("%d/%m/%Y")


@dataclass
class BOMResult:
    success: bool
    xlsx_path: Optional[Path] = None
    pdf_path: Optional[Path] = None
    items: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    error_message: Optional[str] = None


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Material densities in g/cm³ — keys match catalog option values
_DENSITY: dict[str, float] = {
    "ASTM_A36":     7.85,
    "ASTM_A572_G50": 7.85,
    "SS304":         8.00,
    "SS316":         8.00,
    "AL6061T6":      2.70,
}

# Column spec: (header, item_dict_key, xl_width_chars, pdf_width_mm)
_COLUMNS: list[tuple[str, str, int, float]] = [
    ("ÍTEM",            "item_number",     6,   8.0),
    ("N° PARTE",        "part_code",      13,  18.0),
    ("DESCRIPCIÓN",     "description",    38,  60.0),
    ("CANT.",           "quantity",        7,   9.0),
    ("UNIDAD",          "unit",            7,  10.0),
    ("MATERIAL",        "material",       16,  28.0),
    ("NORMA",           "standard",       20,  35.0),
    ("PESO UNIT. (kg)", "unit_weight_kg", 14,  22.0),
    ("OBSERVACIONES",   "observations",   28,  77.0),
]


# ---------------------------------------------------------------------------
# BOMGenerator
# ---------------------------------------------------------------------------

class BOMGenerator:
    """
    Generates a Bill of Materials from piece parameters and catalog template.
    Produces .xlsx + .pdf files; never raises (errors captured in BOMResult).
    """

    def generate(
        self,
        piece_code: str,
        parameters: dict,
        output_dir: Path,
        ctx: BOMContext,
    ) -> BOMResult:
        """Entry point: build items and export both formats."""
        try:
            items = self._build_items(piece_code, parameters, ctx)
            if not items:
                return BOMResult(
                    success=False,
                    error_message=f"Sin plantilla BOM para '{piece_code}'.",
                )

            stem      = f"bom_{ctx.revision_code}"
            warnings: list[str] = []

            xlsx_path = output_dir / f"{stem}.xlsx"
            self._generate_xlsx(items, xlsx_path, ctx)

            pdf_path = output_dir / f"{stem}.pdf"
            self._generate_pdf(items, pdf_path, ctx)

            return BOMResult(
                success=True,
                xlsx_path=xlsx_path,
                pdf_path=pdf_path,
                items=items,
                warnings=warnings,
            )
        except Exception as exc:
            return BOMResult(success=False, error_message=str(exc))

    # ------------------------------------------------------------------
    # Item construction
    # ------------------------------------------------------------------

    def _build_items(
        self, piece_code: str, parameters: dict, ctx: BOMContext
    ) -> list[dict]:
        """Build BOM item dicts from catalog bom_template + runtime parameters."""
        from cad_generator.config.catalog_loader import catalog

        piece_spec = catalog.get_piece(piece_code)
        if not piece_spec or not piece_spec.bom_template:
            return []

        items: list[dict] = []
        for tpl in piece_spec.bom_template:
            # Resolve human-readable material label
            material_label = ""
            if tpl.material_param:
                mat_value = parameters.get(tpl.material_param, "")
                mat_spec  = piece_spec.get_parameter(tpl.material_param)
                if mat_spec:
                    opt = next(
                        (o for o in mat_spec.options if o.value == mat_value), None
                    )
                    material_label = opt.label if opt else str(mat_value)
                else:
                    material_label = str(mat_value)

            # Substitute {drawing_number}, {revision_code}, and any param in observations
            observations = ""
            if tpl.observations:
                fmt_kwargs = dict(parameters)
                fmt_kwargs["drawing_number"] = ctx.drawing_number
                fmt_kwargs["revision_code"]  = ctx.revision_code
                try:
                    observations = tpl.observations.format(**fmt_kwargs)
                except (KeyError, ValueError):
                    observations = tpl.observations  # fallback: verbatim

            unit_weight_kg = self._compute_weight(piece_code, parameters)

            items.append({
                "item_number":    tpl.item_number,
                "part_code":      tpl.part_code,
                "description":    tpl.description,
                "quantity":       float(tpl.quantity),
                "unit":           tpl.unit,
                "material":       material_label,
                "standard":       tpl.standard,
                "unit_weight_kg": unit_weight_kg,
                "observations":   observations,
            })

        return items

    def _compute_weight(self, piece_code: str, parameters: dict) -> Optional[float]:
        """Compute piece weight in kg from volume × material density."""
        if piece_code != "base_plate":
            return None
        density = _DENSITY.get(parameters.get("material", ""))
        if density is None:
            return None
        largo   = float(parameters.get("largo",   0))
        ancho   = float(parameters.get("ancho",   0))
        espesor = float(parameters.get("espesor", 0))
        # mm³ → cm³ (÷1000) → g (×density) → kg (÷1000)
        return largo * ancho * espesor * density / 1_000_000.0

    # ------------------------------------------------------------------
    # Excel export (openpyxl)
    # ------------------------------------------------------------------

    def _generate_xlsx(
        self, items: list[dict], path: Path, ctx: BOMContext
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"

        # Style constants
        _h_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        _b_font  = Font(name="Arial", size=9)
        _t_font  = Font(name="Arial", bold=True, size=9)
        _h_fill  = PatternFill("solid", fgColor="1F4E79")
        _s_fill  = PatternFill("solid", fgColor="D6E4F0")
        _center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _left    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        _thin    = Side(style="thin", color="AAAAAA")
        _border  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

        n_cols = len(_COLUMNS)

        def _merge_row(row: int, value: str, font: Font,
                       align: Alignment = _center) -> None:
            end_col = chr(ord("A") + n_cols - 1)
            ws.merge_cells(f"A{row}:{end_col}{row}")
            cell = ws[f"A{row}"]
            cell.value     = value
            cell.font      = font
            cell.alignment = align

        # Row 1 — company name
        _merge_row(1, ctx.company_name or "—",
                   Font(name="Arial", bold=True, size=14))
        ws.row_dimensions[1].height = 22

        # Row 2 — document title
        _merge_row(2, f"LISTA DE MATERIALES — {ctx.piece_display_name}",
                   Font(name="Arial", bold=True, size=11))
        ws.row_dimensions[2].height = 18

        # Row 3 — metadata cells (no merge)
        meta = [
            ("A3", f"Plano N°: {ctx.drawing_number}"),
            ("B3", f"Rev.: {ctx.revision_code}"),
            ("C3", f"Fecha: {ctx.date_str}"),
            ("D3", f"Dibujó: {ctx.author}"),
            ("E3", "Norma: IRAM 4505"),
        ]
        for addr, val in meta:
            cell = ws[addr]
            cell.value = val
            cell.font  = Font(name="Arial", size=9)
        ws.row_dimensions[3].height = 14

        # Row 4 — blank spacer
        ws.append([])
        ws.row_dimensions[4].height = 6

        # Row 5 — column headers
        ws.append([col[0] for col in _COLUMNS])
        hdr_row = ws.max_row
        for cell in ws[hdr_row]:
            cell.font      = _h_font
            cell.fill      = _h_fill
            cell.alignment = _center
            cell.border    = _border
        ws.row_dimensions[hdr_row].height = 28

        # Data rows
        numeric_keys = {"item_number", "quantity", "unit_weight_kg"}
        for i, item in enumerate(items):
            row_vals = []
            for col in _COLUMNS:
                v = item.get(col[1])
                if isinstance(v, float) and col[1] == "unit_weight_kg":
                    v = round(v, 3) if v is not None else ""
                elif v is None:
                    v = ""
                row_vals.append(v)
            ws.append(row_vals)
            dr = ws.max_row
            alt_fill = _s_fill if i % 2 == 0 else None
            for j, cell in enumerate(ws[dr]):
                cell.font   = _b_font
                cell.border = _border
                if alt_fill:
                    cell.fill = alt_fill
                cell.alignment = _center if _COLUMNS[j][1] in numeric_keys else _left
            ws.row_dimensions[dr].height = 22

        # Total weight row
        total_kg = sum(
            (item.get("unit_weight_kg") or 0.0) * item.get("quantity", 1.0)
            for item in items
            if item.get("unit_weight_kg") is not None
        )
        if total_kg > 0:
            ws.append([])
            peso_col_idx = next(
                (i for i, c in enumerate(_COLUMNS) if c[1] == "unit_weight_kg"), None
            )
            total_row: list = [""] * n_cols
            if peso_col_idx is not None:
                total_row[peso_col_idx - 1] = "PESO TOTAL"
                total_row[peso_col_idx]     = round(total_kg, 3)
                total_row[peso_col_idx + 1] = "kg"
            ws.append(total_row)
            for cell in ws[ws.max_row]:
                cell.font = _t_font

        # Column widths and freeze
        for i, col in enumerate(_COLUMNS):
            ws.column_dimensions[chr(ord("A") + i)].width = col[2]
        ws.freeze_panes = "A6"

        wb.save(str(path))

    # ------------------------------------------------------------------
    # PDF export (reportlab)
    # ------------------------------------------------------------------

    def _generate_pdf(
        self, items: list[dict], path: Path, ctx: BOMContext
    ) -> None:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )

        doc = SimpleDocTemplate(
            str(path),
            pagesize=landscape(A4),
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=15 * mm, bottomMargin=15 * mm,
        )
        styles = getSampleStyleSheet()
        elements = []

        # Header block
        elements.append(
            Paragraph(f"<b>{ctx.company_name or '—'}</b>", styles["Title"])
        )
        elements.append(
            Paragraph(
                f"Lista de Materiales — {ctx.piece_display_name}",
                styles["Heading2"],
            )
        )
        elements.append(
            Paragraph(
                f"Plano N°: {ctx.drawing_number} &nbsp;|&nbsp; "
                f"Rev.: {ctx.revision_code} &nbsp;|&nbsp; "
                f"Fecha: {ctx.date_str} &nbsp;|&nbsp; "
                f"Dibujó: {ctx.author}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 6 * mm))

        # Table
        col_keys     = [c[1] for c in _COLUMNS]
        col_headers  = [c[0] for c in _COLUMNS]
        col_widths   = [c[3] * mm for c in _COLUMNS]

        def _fmt(key: str, val) -> str:
            if val is None:
                return ""
            if key == "quantity":
                v = float(val)
                return f"{v:.0f}" if v == int(v) else f"{v:.2f}"
            if key == "unit_weight_kg":
                return f"{float(val):.3f}"
            return str(val)

        table_data = [col_headers]
        for item in items:
            table_data.append([_fmt(k, item.get(k)) for k in col_keys])

        # Total weight footer row
        total_kg = sum(
            (item.get("unit_weight_kg") or 0.0) * item.get("quantity", 1.0)
            for item in items
            if item.get("unit_weight_kg") is not None
        )
        if total_kg > 0:
            peso_idx = next(
                (i for i, c in enumerate(_COLUMNS) if c[1] == "unit_weight_kg"), None
            )
            foot_row = [""] * len(_COLUMNS)
            if peso_idx is not None:
                foot_row[peso_idx - 1] = "PESO TOTAL"
                foot_row[peso_idx]     = f"{total_kg:.3f}"
                foot_row[peso_idx + 1] = "kg"
            table_data.append(foot_row)

        n_rows = len(table_data)
        _HEADER_BG  = colors.HexColor("#1F4E79")
        _ALT_BG     = colors.HexColor("#D6E4F0")
        _GRID_COLOR = colors.HexColor("#AAAAAA")

        alt_cmds = [
            ("BACKGROUND", (0, i), (-1, i), _ALT_BG)
            for i in range(2, n_rows, 2)
        ]

        tbl = Table(table_data, colWidths=col_widths)
        tbl.setStyle(TableStyle([
            # Header row
            ("BACKGROUND",   (0, 0), (-1, 0), _HEADER_BG),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0), 9),
            ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
            # Data rows
            ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, -1), 8),
            ("ALIGN",        (0, 1), (1, -1),  "CENTER"),   # ÍTEM + CANT. centered
            ("ALIGN",        (2, 1), (-1, -1), "LEFT"),
            # Grid
            ("GRID",         (0, 0), (-1, -1), 0.5, _GRID_COLOR),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            *alt_cmds,
        ]))
        elements.append(tbl)

        doc.build(elements)
