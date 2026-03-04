"""
Tests for BOMGenerator — Semana 11.

Covers:
  - build_items: correct count, material label resolution, observation substitution
  - Weight computation for base_plate
  - Excel file creation and basic content checks
  - PDF file creation
  - BOMContext date auto-fill
  - Unsupported piece_code returns failure
"""

from __future__ import annotations

from pathlib import Path

import pytest

from cad_generator.core.bom_generator import BOMContext, BOMGenerator, BOMResult


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

_PARAMS: dict = {
    "largo":  300.0,
    "ancho":  200.0,
    "espesor": 12.0,
    "material": "ASTM_A36",
    "patron_perforaciones": "rectangular_4",
    "diametro_perforacion": 18.0,
    "margen_perforacion":   30.0,
    "tiene_ranuras": False,
    "ancho_ranura":  12.0,
    "largo_ranura":  40.0,
}


def _ctx(**overrides) -> BOMContext:
    defaults = dict(
        piece_code="base_plate",
        piece_display_name="Placa Base Estructural",
        drawing_number="PL-001",
        revision_code="A",
        author="Fede",
        company_name="TestCo",
    )
    defaults.update(overrides)
    return BOMContext(**defaults)


# ---------------------------------------------------------------------------
# BOMContext
# ---------------------------------------------------------------------------

class TestBOMContext:

    def test_date_str_auto_filled(self):
        ctx = _ctx()
        assert ctx.date_str != ""
        parts = ctx.date_str.split("/")
        assert len(parts) == 3
        assert all(p.isdigit() for p in parts)

    def test_custom_date_preserved(self):
        ctx = _ctx(date_str="01/06/2026")
        assert ctx.date_str == "01/06/2026"


# ---------------------------------------------------------------------------
# build_items
# ---------------------------------------------------------------------------

class TestBuildItems:

    def _build(self, params=None, **ctx_overrides):
        gen = BOMGenerator()
        ctx = _ctx(**ctx_overrides)
        return gen._build_items("base_plate", params or _PARAMS, ctx)

    def test_base_plate_returns_one_item(self):
        items = self._build()
        assert len(items) == 1

    def test_item_has_required_keys(self):
        item = self._build()[0]
        for key in ("item_number", "part_code", "description",
                    "quantity", "unit", "material", "standard", "observations"):
            assert key in item

    def test_material_label_resolved(self):
        """'ASTM_A36' in params → human-readable label in item.material."""
        item = self._build()[0]
        # Label should NOT be the raw catalog value
        assert item["material"] != "ASTM_A36"
        assert "A36" in item["material"]  # e.g. "Acero ASTM A36"

    def test_observation_substitution(self):
        item = self._build(drawing_number="PL-999", revision_code="B")[0]
        assert "PL-999" in item["observations"]
        assert "B" in item["observations"]

    def test_unit_weight_computed_for_astm_a36(self):
        """300×200×12 × 7.85 g/cm³ / 1e6 ≈ 5.652 kg."""
        item = self._build()[0]
        w = item["unit_weight_kg"]
        assert w is not None
        assert pytest.approx(w, rel=1e-3) == 300 * 200 * 12 * 7.85 / 1_000_000

    def test_unit_weight_none_for_unknown_material(self):
        params = {**_PARAMS, "material": "UNKNOWN_MAT"}
        items = self._build(params=params)
        assert items[0]["unit_weight_kg"] is None

    def test_aluminum_density(self):
        """AL6061T6 density = 2.70 g/cm³."""
        params = {**_PARAMS, "material": "AL6061T6"}
        item = self._build(params=params)[0]
        expected = 300 * 200 * 12 * 2.70 / 1_000_000
        assert pytest.approx(item["unit_weight_kg"], rel=1e-3) == expected

    def test_unsupported_piece_returns_empty(self):
        gen = BOMGenerator()
        items = gen._build_items("unknown_piece", _PARAMS, _ctx())
        assert items == []


# ---------------------------------------------------------------------------
# Weight computation
# ---------------------------------------------------------------------------

class TestComputeWeight:

    def _gen(self):
        return BOMGenerator()

    def test_astm_a36_weight(self):
        w = self._gen()._compute_weight("base_plate", _PARAMS)
        assert w is not None
        assert w > 0

    def test_unknown_material_returns_none(self):
        params = {**_PARAMS, "material": "UNKNOWN"}
        assert self._gen()._compute_weight("base_plate", params) is None

    def test_non_plate_piece_returns_none(self):
        assert self._gen()._compute_weight("other_piece", _PARAMS) is None

    def test_ss304_density(self):
        params = {**_PARAMS, "material": "SS304"}
        w = self._gen()._compute_weight("base_plate", params)
        expected = 300 * 200 * 12 * 8.00 / 1_000_000
        assert pytest.approx(w, rel=1e-3) == expected


# ---------------------------------------------------------------------------
# File generation (integration)
# ---------------------------------------------------------------------------

class TestBOMGeneration:

    def test_generate_returns_success(self, tmp_path):
        result = BOMGenerator().generate(
            piece_code="base_plate",
            parameters=_PARAMS,
            output_dir=tmp_path,
            ctx=_ctx(revision_code="A"),
        )
        assert result.success

    def test_xlsx_file_created(self, tmp_path):
        result = BOMGenerator().generate(
            piece_code="base_plate",
            parameters=_PARAMS,
            output_dir=tmp_path,
            ctx=_ctx(revision_code="A"),
        )
        assert result.xlsx_path is not None
        assert result.xlsx_path.exists()
        assert result.xlsx_path.suffix == ".xlsx"

    def test_pdf_file_created(self, tmp_path):
        result = BOMGenerator().generate(
            piece_code="base_plate",
            parameters=_PARAMS,
            output_dir=tmp_path,
            ctx=_ctx(revision_code="A"),
        )
        assert result.pdf_path is not None
        assert result.pdf_path.exists()
        assert result.pdf_path.suffix == ".pdf"

    def test_revision_code_in_filenames(self, tmp_path):
        result = BOMGenerator().generate(
            piece_code="base_plate",
            parameters=_PARAMS,
            output_dir=tmp_path,
            ctx=_ctx(revision_code="C"),
        )
        assert "C" in result.xlsx_path.name
        assert "C" in result.pdf_path.name

    def test_items_returned(self, tmp_path):
        result = BOMGenerator().generate(
            piece_code="base_plate",
            parameters=_PARAMS,
            output_dir=tmp_path,
            ctx=_ctx(revision_code="A"),
        )
        assert len(result.items) >= 1

    def test_unsupported_piece_returns_failure(self, tmp_path):
        result = BOMGenerator().generate(
            piece_code="unknown_piece_xyz",
            parameters=_PARAMS,
            output_dir=tmp_path,
            ctx=_ctx(piece_code="unknown_piece_xyz"),
        )
        assert not result.success
        assert result.error_message is not None

    def test_xlsx_contains_data(self, tmp_path):
        """Open the generated xlsx and verify the header row is present."""
        from openpyxl import load_workbook

        result = BOMGenerator().generate(
            piece_code="base_plate",
            parameters=_PARAMS,
            output_dir=tmp_path,
            ctx=_ctx(revision_code="A"),
        )
        wb = load_workbook(str(result.xlsx_path))
        ws = wb.active
        # Row 1 should contain company name
        assert ws["A1"].value is not None
        # Row 5 should be the header row — check first column
        header_cell = ws["A5"].value
        assert header_cell is not None and "ÍTEM" in str(header_cell)

    def test_no_warnings_for_valid_params(self, tmp_path):
        result = BOMGenerator().generate(
            piece_code="base_plate",
            parameters=_PARAMS,
            output_dir=tmp_path,
            ctx=_ctx(revision_code="A"),
        )
        assert result.success
        assert result.warnings == []
